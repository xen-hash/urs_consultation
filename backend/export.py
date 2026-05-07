import io
import pytz
from datetime import datetime
from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import query
from config import PROFESSOR_LIST

export_bp = Blueprint("export", __name__)
PH = pytz.timezone("Asia/Manila")

HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_row(ws, row_num, num_cols, alt=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


@export_bp.route("/export", methods=["GET"])
def export_data():
    export_type = request.args.get("type", "today")
    now_ph = datetime.now(PH)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary Sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "URS Faculty Consultation System — Export"
    ws_sum["A1"].font = Font(bold=True, size=14, color="003366")
    ws_sum["A2"] = f"Generated: {now_ph.strftime('%B %d, %Y %I:%M %p')} PHT"
    ws_sum["A3"] = f"Export Type: {'Today (' + now_ph.strftime('%Y-%m-%d') + ')' if export_type == 'today' else 'All Records'}"

    # Dept summary
    ws_sum["A5"] = "Department"
    ws_sum["B5"] = "Total Requests"
    ws_sum["C5"] = "Pending"
    ws_sum["D5"] = "Done"
    ws_sum["E5"] = "Declined"
    _style_header(ws_sum, ["Department", "Total Requests", "Pending", "Done", "Declined"], row=5)

    sum_row = 6
    for dept in PROFESSOR_LIST:
        if export_type == "today":
            today_str = now_ph.strftime("%Y-%m-%d")
            rows = query(
                "SELECT status FROM consultation_requests WHERE department=%s AND DATE(created_at)=%s",
                (dept, today_str), fetchall=True
            )
        else:
            rows = query(
                "SELECT status FROM consultation_requests WHERE department=%s",
                (dept,), fetchall=True
            )
        total = len(rows)
        pending  = sum(1 for r in rows if r["status"] == "pending")
        done     = sum(1 for r in rows if r["status"] == "done")
        declined = sum(1 for r in rows if r["status"] == "declined")
        for col, val in enumerate([dept, total, pending, done, declined], 1):
            ws_sum.cell(row=sum_row, column=col, value=val)
        _style_row(ws_sum, sum_row, 5, alt=sum_row % 2 == 0)
        sum_row += 1
    _auto_width(ws_sum)

    # ── Per-Department Consultation Sheet ──────────────────────────────────────
    for dept, profs in PROFESSOR_LIST.items():
        safe_name = dept[:28].replace("/", "-")
        ws = wb.create_sheet(f"CR-{safe_name[:20]}")
        headers = ["ID", "Student ID", "Student Name", "Course", "Professor",
                   "Purpose", "Category", "Status", "Request Time"]
        _style_header(ws, headers)
        ws.row_dimensions[1].height = 20

        if export_type == "today":
            today_str = now_ph.strftime("%Y-%m-%d")
            rows = query(
                """SELECT * FROM consultation_requests
                   WHERE department=%s AND DATE(created_at)=%s
                   ORDER BY created_at DESC""",
                (dept, today_str), fetchall=True
            )
        else:
            rows = query(
                "SELECT * FROM consultation_requests WHERE department=%s ORDER BY created_at DESC",
                (dept,), fetchall=True
            )

        for r_idx, row in enumerate(rows or [], start=2):
            vals = [
                row["id"], row["student_id"], row["student_name"], row["course"],
                row["professor_name"], row["purpose"], row["category"],
                row["status"],
                str(row["request_time"]) if row["request_time"] else ""
            ]
            for c_idx, val in enumerate(vals, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _style_row(ws, r_idx, len(headers), alt=r_idx % 2 == 0)
        _auto_width(ws)

    # ── Teacher Logs Sheet ─────────────────────────────────────────────────────
    ws_tl = wb.create_sheet("Teacher Logs")
    tl_headers = ["ID", "Professor", "Department", "Action", "Manual Status",
                  "Manual Override", "Log Time"]
    _style_header(ws_tl, tl_headers)
    if export_type == "today":
        today_str = now_ph.strftime("%Y-%m-%d")
        tl_rows = query(
            "SELECT * FROM teacher_logs WHERE DATE(created_at)=%s ORDER BY created_at DESC",
            (today_str,), fetchall=True
        )
    else:
        tl_rows = query("SELECT * FROM teacher_logs ORDER BY created_at DESC", fetchall=True)

    for r_idx, row in enumerate(tl_rows or [], start=2):
        vals = [
            row["id"], row["professor_name"], row["department"],
            row["action_type"], row["manual_status"],
            "Yes" if row["manual"] else "No",
            str(row["log_time"]) if row["log_time"] else ""
        ]
        for c_idx, val in enumerate(vals, 1):
            ws_tl.cell(row=r_idx, column=c_idx, value=val)
        _style_row(ws_tl, r_idx, len(tl_headers), alt=r_idx % 2 == 0)
    _auto_width(ws_tl)

    # Save to buffer and send
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"URS_Consultation_{export_type}_{now_ph.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
